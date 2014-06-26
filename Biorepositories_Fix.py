# Clean up spreadsheet to fit C4P format
from openpyxl import load_workbook, cell, Workbook
from openpyxl.cell import coordinate_from_string, column_index_from_string
import requests
from datetime import datetime, date, time

wb = load_workbook('biorepositories_from_website.xlsx')
ws = wb.get_active_sheet()
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "GRbio repos"

START_ROW = 2
END_ROW = ws.get_highest_row()
GR_URL_COL = 'A'
GR_INSTIT_NAME_COL = 'B'
GR_REPO_NAME_COL = 'C'
GR_STATUS_COL = 'H'

#class Entry(self, name):
#	self.name = name
	#institutionName
	#repoName
	#URL
	#desc
	#contact

for row in ws.range('%s%s:%s%s'%(GR_REPO_NAME_COL, START_ROW, GR_REPO_NAME_COL, END_ROW)):
	for cell in row:
		coordinate = coordinate_from_string(cell.get_coordinate())
		if cell.value == None or cell.value == "Herbarium" or cell.value == "Herbario":
			cell.value = ws['%s%s'%(GR_INSTIT_NAME_COL, coordinate[1])].value

wb.save("biorepositories_from_website.xlsx")